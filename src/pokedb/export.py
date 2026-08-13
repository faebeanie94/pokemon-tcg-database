"""Write the Excel workbook (and optional CSVs) from the built database.

The workbook is the deliverable: one file holding every set and card across
games and languages, with an autofilter on each sheet so a card can be found
by game, set name, number or card name. It is regenerated from scratch on
every update.
"""

from __future__ import annotations

import csv
import gzip
import sqlite3
from pathlib import Path

from .config import DB_PATH, EXPORTS

WORKBOOK_NAME = "Card_Database.xlsx"
CARDS_BY_GAME_NAME = "cards_by_game.xlsx"  # one sheet per game + language

CARD_QUERY = """
SELECT
    game_name            AS "Game",
    language_name        AS "Language",
    set_name             AS "Set Name",
    CASE
      -- Pokémon / TCG printed form: 4/102, 001/198, TG12/198
      WHEN game <> 'sports'
           AND card_count IS NOT NULL
           AND instr(card_number, '/') = 0
      THEN card_number || '/' || card_count
      ELSE card_number
    END                  AS "Card Number",
    COALESCE(display_name, card_name) AS "Card Name",
    subject_name         AS "Subject",
    parallel             AS "Parallel",
    notations            AS "Notations",
    CASE
      WHEN serial_number IS NOT NULL AND print_run IS NOT NULL
      THEN serial_number || '/' || print_run
    END                  AS "Serial",
    manufacturer         AS "Manufacturer",
    sport                AS "Sport",
    product_year         AS "Season",
    release_year         AS "Year",
    set_name_en          AS "Set Name (EN)",
    card_name_en         AS "Card Name (EN)",
    set_abbreviation     AS "Set Code",
    release_date         AS "Release Date",
    card_count           AS "Cards In Set"
FROM (
    SELECT c.game, g.name AS game_name, c.language, l.name_en AS language_name,
           s.name AS set_name, s.name_en AS set_name_en,
           s.abbreviation AS set_abbreviation, s.release_date, s.release_year,
           s.manufacturer, s.sport, s.product_year,
           c.number AS card_number, c.number_prefix, c.number_value,
           c.name AS card_name, c.name_en AS card_name_en,
           c.subject_name, c.parallel, c.notations,
           c.serial_number, c.print_run, c.display_name,
           -- Prefer the printed set size (official) so Base Set is /102 not loaded count.
           s.card_count_official AS card_count
    FROM cards c
    JOIN sets s      ON s.set_uid = c.set_uid
    JOIN games g     ON g.code = c.game
    JOIN languages l ON l.code = c.language
)
ORDER BY game, language, COALESCE(release_date, '9999'), set_name,
         COALESCE(number_prefix, ''), COALESCE(number_value, 999999), card_number
"""

SET_QUERY = """
SELECT
    g.name                       AS "Game",
    l.name_en                    AS "Language",
    s.name                       AS "Set Name",
    s.name_en                    AS "Set Name (EN)",
    s.abbreviation               AS "Set Code",
    s.manufacturer               AS "Manufacturer",
    s.sport                      AS "Sport",
    s.product_year               AS "Season",
    s.release_year               AS "Year",
    s.release_date               AS "Release Date",
    s.series_name                AS "Series",
    s.card_count_official        AS "Cards In Set",
    s.card_count_loaded          AS "Cards Listed",
    s.sources                    AS "Sources"
FROM sets s
JOIN games g     ON g.code = s.game
JOIN languages l ON l.code = s.language
ORDER BY s.game, s.language, COALESCE(s.release_date, '9999'), s.name
"""

COVERAGE_QUERY = """
SELECT language_name AS "Language", language AS "Code", sets AS "Sets",
       sets_with_cards AS "Sets With Cards", cards AS "Cards",
       first_release AS "First Release", latest_release AS "Latest Release"
FROM coverage_by_language
ORDER BY cards DESC, sets DESC
"""

GAME_COVERAGE_QUERY = """
SELECT game_name AS "Game", game AS "Code", game_kind AS "Kind",
       sets AS "Sets", cards AS "Cards"
FROM coverage_by_game
ORDER BY cards DESC, sets DESC
"""


def _connect(db_path: Path) -> sqlite3.Connection:
    if not db_path.exists():
        raise SystemExit(f"{db_path} not found - run `python -m pokedb build` first")
    return sqlite3.connect(db_path)


def _fetch(connection: sqlite3.Connection, query: str) -> tuple[list[str], list[tuple]]:
    cursor = connection.execute(query)
    return [column[0] for column in cursor.description], cursor.fetchall()


def export_all(db_path: Path = DB_PATH, write_csv: bool = True) -> list[Path]:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    connection = _connect(db_path)

    card_headers, card_rows = _fetch(connection, CARD_QUERY)
    set_headers, set_rows = _fetch(connection, SET_QUERY)
    coverage_headers, coverage_rows = _fetch(connection, COVERAGE_QUERY)
    game_headers, game_rows = _fetch(connection, GAME_COVERAGE_QUERY)
    built_at = connection.execute(
        "SELECT value FROM build_info WHERE key = 'built_at'"
    ).fetchone()
    sources = connection.execute(
        "SELECT value FROM build_info WHERE key = 'source_order'"
    ).fetchone()
    connection.close()

    written = [
        _write_cards_by_game(card_headers, card_rows, game_headers, game_rows),
        _write_workbook(
            card_headers,
            card_rows,
            set_headers,
            set_rows,
            coverage_headers,
            coverage_rows,
            game_headers,
            game_rows,
            built_at[0] if built_at else "",
            sources[0] if sources else "",
        ),
    ]

    if write_csv:
        written.append(_write_csv(set_headers, set_rows, EXPORTS / "sets.csv"))
        written.append(
            _write_csv(card_headers, card_rows, EXPORTS / "cards.csv.gz", compress=True)
        )
    return written


def _write_csv(headers: list[str], rows, path: Path, compress: bool = False) -> Path:
    opener = (
        (lambda: gzip.open(path, "wt", newline="", encoding="utf-8"))
        if compress
        else (lambda: path.open("w", newline="", encoding="utf-8-sig"))
    )
    with opener() as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def _sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet titles: max 31 chars, no ``: \\ / ? * [ ]``."""
    cleaned = "".join("-" if ch in r"[]:*?/\\" else ch for ch in name)[:31].strip() or "Game"
    base = cleaned
    counter = 2
    while cleaned in used:
        suffix = f"_{counter}"
        cleaned = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(cleaned)
    return cleaned


def _write_cards_by_game(
    card_headers: list[str],
    card_rows: list[tuple],
    game_headers: list[str],
    game_rows: list[tuple],
) -> Path:
    """One worksheet per game + language (e.g. ``Pokemon EN``, ``MTG Japanese``)."""
    from collections import defaultdict

    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font

    path = EXPORTS / CARDS_BY_GAME_NAME
    # (game_name, language_name) -> rows
    buckets: dict[tuple[str, str], list[tuple]] = defaultdict(list)
    for row in card_rows:
        buckets[(str(row[0]), str(row[1]))].append(row)

    # Order games by coverage (largest first), languages by row count within game.
    game_rank = {str(row[0]): index for index, row in enumerate(game_rows)}
    counts: dict[tuple[str, str], int] = {key: len(rows) for key, rows in buckets.items()}

    def sort_key(item: tuple[tuple[str, str], list[tuple]]) -> tuple:
        (game, language), rows = item
        return (game_rank.get(game, 10_000), -len(rows), language)

    ordered = sorted(buckets.items(), key=sort_key)

    # Short labels so Excel's 31-char sheet limit stays readable.
    game_abbrev = {
        "Pokémon TCG": "Pokemon",
        "Magic: The Gathering": "MTG",
        "Yu-Gi-Oh!": "YGO",
        "One Piece Card Game": "OnePiece",
        "Disney Lorcana": "Lorcana",
        "Flesh and Blood": "FAB",
        "Weiss Schwarz": "Weiss",
        "Dragon Ball Z TCG": "DBZ",
        "Dragon Ball Super: Masters": "DBS",
        "Dragon Ball Super: Fusion World": "DBSFW",
        "Marvel Dice Masters": "DiceMast",
        "Warhammer Age of Sigmar Champions": "AoSChamp",
        "Sports & Entertainment Cards": "Sports",
        "MetaZoo": "MetaZoo",
        "UniVersus": "UniVersus",
    }
    lang_abbrev = {
        "English": "EN",
        "French": "FR",
        "German": "DE",
        "Spanish": "ES",
        "Italian": "IT",
        "Portuguese": "PT",
        "Portuguese (Brazil)": "PT-BR",
        "Japanese": "JA",
        "Korean": "KO",
        "Chinese (Traditional)": "ZH-TW",
        "Chinese (Simplified)": "ZH-CN",
        "Chinese (Simplified, MTG)": "ZHS",
        "Chinese (Traditional, MTG)": "ZHT",
        "Indonesian": "ID",
        "Thai": "TH",
        "Dutch": "NL",
        "Polish": "PL",
        "Russian": "RU",
        "Undetermined": "UND",
    }

    workbook = Workbook(write_only=True)
    used: set[str] = set()
    for (game, language), rows in ordered:
        g = game_abbrev.get(game, game)
        lang = lang_abbrev.get(language, language)
        title = _sheet_title(f"{g} {lang}", used)
        sheet = workbook.create_sheet(title)
        header_cells = []
        for header in card_headers:
            cell = WriteOnlyCell(sheet, value=header)
            cell.font = Font(bold=True)
            header_cells.append(cell)
        sheet.append(header_cells)
        for row in rows:
            sheet.append(list(row))

    workbook.save(path)
    return path


def _write_workbook(
    card_headers,
    card_rows,
    set_headers,
    set_rows,
    coverage_headers,
    coverage_rows,
    game_headers,
    game_rows,
    built_at: str,
    sources: str,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    path = EXPORTS / WORKBOOK_NAME
    workbook = Workbook()
    workbook.remove(workbook.active)

    widths = {
        "Game": 22, "Language": 20, "Set Name": 42, "Set Name (EN)": 38,
        "Card Name": 40, "Card Name (EN)": 30, "Card Number": 14, "Year": 8,
        "Set Code": 12, "Release Date": 14, "Cards In Set": 13, "Cards Listed": 13,
        "Series": 26, "Sources": 30, "Code": 12, "Sets": 8, "Cards": 10,
        "Sets With Cards": 16, "First Release": 14, "Latest Release": 14,
        "Subject": 28, "Parallel": 18, "Notations": 14, "Serial": 12,
        "Manufacturer": 14, "Sport": 12, "Season": 12, "Kind": 10,
    }

    def add_sheet(title: str, headers: list[str], rows) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(list(row))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        if rows:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 16)

    add_sheet("Cards", card_headers, card_rows)
    add_sheet("Sets", set_headers, set_rows)
    add_sheet("Coverage", coverage_headers, coverage_rows)
    add_sheet("By Game", game_headers, game_rows)

    about = workbook.create_sheet("About")
    about.column_dimensions["A"].width = 26
    about.column_dimensions["B"].width = 96
    notes = [
        ("Multi-game Card Database", ""),
        ("Generated", built_at),
        ("Sources", sources),
        ("Cards", f"{len(card_rows):,} rows"),
        ("Sets", f"{len(set_rows):,} rows"),
        ("", ""),
        (
            "Identity",
            "card_uid is '<game>:<language>:<set>#<number>' and appends '#<parallel>' "
            "when a parallel/variant is present.",
        ),
        (
            "Sports grading",
            "Set name + card name/parallel + number. Serials like 09/15 are print runs, "
            "not Pokémon-style printed totals.",
        ),
        (
            "Updating",
            "Run `python -m pokedb update` to rebuild. Fetch other sources with "
            "`python -m pokedb fetch --source tcgcsv --game onepiece` etc.",
        ),
        (
            "Licensing",
            "Internal grading use only unless you hold redistribution rights from "
            "Bandai/Konami/Wizards/Topps/Panini/etc.",
        ),
    ]
    for label, value in notes:
        about.append([label, value])
    about["A1"].font = Font(bold=True, size=14)
    for row in about.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].font = Font(bold=True)
    for row in about.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)
    return path
